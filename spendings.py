from json import load
import datetime
from datetime import date
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl import load_workbook

wb = load_workbook(filename = "spendings.xlsx")
ws = wb["Document"]
daySinceSept4 = (date.today() - date(2022, 9, 4)).days

r = 2
tempSum = 0
while True:
    c1 = ws.cell(row = r, column = 1).value
    c2 = ws.cell(row = r + 1, column = 1).value
    if(ws.cell(row = r + 1, column = 3).value == 1):
        amt = ws.cell(row = r, column = 2).value
    if(c2 == None):
        break
    elif(str(c1)[:10] != str(c2)[:10]):
        ws.cell(row = r, column = 8).value = str(c1)[:10]
        tempSum += amt
        ws.cell(row = r, column = 9).value = tempSum
        tempSum = 0
    else:
        ws.cell(row = r, column = 9).value = None
        tempSum += amt
    r += 1
wb.save(filename = "spendings.xlsx")
